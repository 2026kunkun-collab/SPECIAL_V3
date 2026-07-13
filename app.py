import os
import io
import json
import time
import re
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from google import genai
from google.genai import types
from pydantic import BaseModel, Field

app = Flask(__name__)

# 📋 STRICT PYDANTIC STRUCTURED SCHEMA FOR THE AI EXTRACTOR
class ConcreteDeliverySchema(BaseModel):
    company_name: str = Field(description="The exact text found at the top left corner representing the company, e.g., 'PERFECT READYMIX (SELATAN) SDN. BHD.'")
    do_number: str = Field(description="The D/O No. found at the top right, remove all spaces, e.g., 'PJ006962'")
    date_str: str = Field(description="The date field exactly as written, e.g., '30/06/2026'")
    time_dispatched: str = Field(description="The Time Dispatched value, e.g., '15:14:18' or '15:41:24'")
    truck_no: str = Field(description="The Truck No. value, e.g., 'JPP5396' or 'BMX7169'")
    driver_name: str = Field(description="The Driver's Name found at the bottom middle center, e.g., 'ANDI' or 'MANSUR'")
    customer: str = Field(description="The text in the Customer box, e.g., 'TSB BUILDERS SDN BHD'")
    delivered_to: str = Field(description="The text in the Delivered To box, e.g., 'PERMAS JAYA'")
    this_load: float = Field(description="The numeric volume of This Load (M³), e.g., 8.00")

def format_excel_date(date_str):
    """Converts varying date inputs to strict DD.MM.YY format"""
    if not date_str:
        return ""
    date_clean = str(date_str).strip().replace('.', '/')
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(date_clean, fmt)
            return dt.strftime("%d.%m.%y")
        except ValueError:
            continue
    return date_clean

def format_excel_month(date_str):
    """Extracts short month notation and year, e.g., Jun-26"""
    if not date_str:
        return ""
    date_clean = str(date_str).strip().replace('.', '/')
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(date_clean, fmt)
            return dt.strftime("%b-%y")
        except ValueError:
            continue
    return ""

def format_excel_time(time_str):
    """Converts 24h standard military string (15:14:18) to 12h AM/PM format (3:14PM)"""
    if not time_str:
        return ""
    time_clean = str(time_str).strip()
    # Match HH:MM or HH:MM:SS
    match = re.match(r'^(\d{1,2}):(\d{2})', time_clean)
    if match:
        hours = int(match.group(1))
        minutes = match.group(2)
        suffix = "AM"
        if hours >= 12:
            suffix = "PM"
            if hours > 12:
                hours -= 12
        elif hours == 0:
            hours = 12
        return f"{hours}:{minutes}{suffix}"
    return time_clean

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scan-paper', methods=['POST'])
def scan_paper():
    if 'paper_image' not in request.files:
        return jsonify({"status": "error", "message": "No image uploaded"}), 400
        
    image_file = request.files['paper_image']
    image_bytes = image_file.read()
    user_api_key = request.form.get('user_api_key', '').strip()
    active_api_key = user_api_key or os.environ.get("GEMINI_API_KEY")
    
    if not active_api_key:
        return jsonify({"status": "error", "message": "Missing Gemini API Key."}), 401
    
    prompt = "Analyze this Concrete Ready-Mix Delivery Order receipt image and extract details accurately according to the structural schema definitions."
    
    retries = 3
    delay = 1
    for i in range(retries):
        try:
            temp_client = genai.Client(api_key=active_api_key)
            response = temp_client.models.generate_content(
                model='gemini-2.5-flash',
                contents=[
                    types.Part.from_bytes(data=image_bytes, mime_type=image_file.mimetype),
                    prompt
                ],
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=ConcreteDeliverySchema,
                ),
            )
            extracted_data = json.loads(response.text)
            
            # Format and normalize data fields on the backend
            processed_data = {
                "month": format_excel_month(extracted_data.get("date_str")),
                "date": format_excel_date(extracted_data.get("date_str")),
                "ac_name": f"{extracted_data.get('company_name', '').strip()} - {extracted_data.get('delivered_to', '').strip()}".upper().replace("..", "."),
                "do_number": extracted_data.get("do_number", "").replace(" ", "").upper(),
                "time_dispatched": format_excel_time(extracted_data.get("time_dispatched")),
                "lorry_no": extracted_data.get("truck_no", "").replace(" ", "").upper(),
                "driver": extracted_data.get("driver_name", "").strip().upper(),
                "buyer": extracted_data.get("customer", "").strip().upper(),
                "delivered": extracted_data.get("delivered_to", "").strip().upper(),
                "qty": extracted_data.get("this_load", 0.0)
            }
            return jsonify({"status": "success", "data": processed_data})
            
        except Exception as e:
            if i == retries - 1:
                return jsonify({"status": "error", "message": f"AI Parsing Error: {str(e)}"}), 500
            time.sleep(delay)
            delay *= 2

@app.route('/submit', methods=['POST'])
def submit_data():
    file = request.files.get('excel_file')
    entries = json.loads(request.form.get('entries', '[]'))

    if file and file.filename != '':
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        filename = file.filename
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        filename = "delivery_log.xlsx"

    all_rows = []
    has_header = False
    
    # Read preexisting data, avoiding blank filler spacer cells
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None or str(row[0]).strip().lower() in ("mont", "month"):
            if row and str(row[0]).strip().lower() in ("mont", "month"):
                has_header = True
            continue
        # Convert row tuple into a structured list padding up to 14 elements
        row_pad = list(row) + [""] * (14 - len(row))
        all_rows.append(row_pad[:14])

    # Append fresh incoming web client inputs
    for entry in entries:
        # Save Nota/Pemandu value as text explicitly with leading apostrophe format
        nota_val = str(entry.get('nota', '')).strip()
        
        row_data = [
            entry.get('month', ''),
            entry.get('date', ''),
            "",  # Column C (A/C) is left entirely blank
            entry.get('ac_name', ''),
            entry.get('do_number', ''),
            entry.get('time_dispatched', ''),
            "",  # Column G is left blank
            nota_val,
            entry.get('lorry_no', ''),
            entry.get('driver', ''),
            entry.get('buyer', ''),
            entry.get('delivered', ''),
            float(entry.get('qty', 0)),
            float(entry.get('qty', 0))  # Column N matches Column M exactly
        ]
        all_rows.append(row_data)

    # 🎯 ALPHANUMERIC SEQUENCE SORTING BY D/O NUMBER FROM SMALLEST TO LARGEST
    def extract_do_numeric_key(row_item):
        do_val = str(row_item[4]).strip()
        # Extract digits out of alphanumeric strings (like PJ006962 -> 6962)
        digits = re.findall(r'\d+', do_val)
        return int(digits[0]) if digits else do_val

    try:
        all_rows.sort(key=extract_do_numeric_key)
    except Exception:
        all_rows.sort(key=lambda x: str(x[4])) # Fallback raw string sort

    # Clear worksheet and write structural layout clean
    ws.delete_rows(1, ws.max_row)
    
    headers = [
        "Mont", "Date", "A/C", "A/C Name", "D/O No", 
        "Time Dispatched", "Nota", "Pemandu", "Lorry No", 
        "Driver", "BUYER", "DELIVERED", "Qty (M.cu)", "Actual Qty"
    ]
    ws.append(headers)
    
    # Format number displays for decimal properties explicitly inside Excel cells
    for r_idx, row_data in enumerate(all_rows, start=2):
        ws.append(row_data)
        # Ensure column M and N render as explicit float formatted cells
        ws.cell(row=r_idx, column=13).number_format = '0.00'
        ws.cell(row=r_idx, column=14).number_format = '0.00'

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    return send_file(
        excel_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)